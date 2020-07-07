import openpyxl

wb = openpyxl.load_workbook(r'../files/Data IHM.xlsx')
sheet = wb.active
max_col = sheet.max_column
min_col = sheet.min_column
max_row = sheet.max_row
min_row = sheet.min_row


def getNumColName(name_col):
    for i in range(min_col, max_col + 1):
        v = sheet.cell(row=min_row, column=i)
        L = str(v.value)
        if L == name_col:
            return i


def getNumColLongitude(name_col):
    for i in range(min_col, max_col + 1):
        v = sheet.cell(row=min_row, column=i)
        L = str(v.value)
        if L == name_col:
            return i


def getNumColLatitude(name_col):
    for i in range(min_col, max_col + 1):
        v = sheet.cell(row=min_row, column=i)
        l = str(v.value)
        if l == name_col:
            return i
