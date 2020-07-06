import datetime
import time
from functools import partial
from threading import Event

import openpyxl
from PyQt5.QtCore import QTimer, Qt
from PyQt5.QtGui import QPixmap
from PyQt5.QtWebEngineWidgets import QWebEngineView
from PyQt5.QtWidgets import QVBoxLayout, QHBoxLayout, QTableWidget, QGridLayout, QWidget, QLabel, \
    QApplication, QPushButton, QHeaderView, QLCDNumber, QSpinBox, QTableWidgetItem
from PyQt5 import QtCore, QtGui, QtWidgets

from controllers.timeCtrl import Controller
from gui.analog import AnalogGaugeWidget


class Dragline(object):

    def __init__(self, data):
        """Controller initializer."""
        self.dataConducteur = data

    def setupUi(self, MainWindow):

        self.poste = self.dataConducteur[3]

        # MainWindow.resize(860, 600)

        self.vBox = QVBoxLayout(MainWindow)
        self.vBox.setSpacing(0)
        self.vBox.setContentsMargins(0, 0, 0, 0)

        """header title"""

        self.titleWidget = QWidget()
        self.titleWidget.setObjectName("titleWidget")

        self.hBox = QHBoxLayout()
        self.titleWidget.setLayout(self.hBox)

        self.vBox.addWidget(self.titleWidget, 5)

        self.imgOCP = QLabel()
        self.pixmapOCP = QPixmap('../images/logo ocp.png')
        self.pixmapSizeOCP = self.pixmapOCP.scaled(100, 100, Qt.KeepAspectRatio, Qt.FastTransformation)
        self.imgOCP.setPixmap(self.pixmapSizeOCP)
        self.imgOCP.setAlignment(QtCore.Qt.AlignLeft)

        self.hBox.addWidget(self.imgOCP)

        self.lblBigTitle = QLabel()
        self.lblBigTitle.setText("Dragline Dashboard")
        self.lblBigTitle.setAlignment(QtCore.Qt.AlignCenter)
        self.lblBigTitle.setObjectName("lblBigTitle")

        self.hBox.addWidget(self.lblBigTitle)

        self.imgPillioty = QLabel()
        self.pixmapPillioty = QPixmap('../images/pillioty logo.png')
        self.pixmapSizePillioty = self.pixmapPillioty.scaled(100, 100, Qt.KeepAspectRatio, Qt.FastTransformation)
        self.imgPillioty.setAlignment(QtCore.Qt.AlignRight)
        self.imgPillioty.setPixmap(self.pixmapSizePillioty)

        self.hBox.addWidget(self.imgPillioty)

        """header dragline information"""

        self.infoWidget = QWidget()
        self.infoWidget.setObjectName("infowidget")

        self.hBoxDraglineInfo = QHBoxLayout()
        self.hBoxDraglineInfo.setSpacing(10)
        self.infoWidget.setLayout(self.hBoxDraglineInfo)

        self.vBox.addWidget(self.infoWidget, 17)

        self.gridLayout1 = QGridLayout()
        self.hBoxDraglineInfo.addLayout(self.gridLayout1)

        self.lblDraglineId = QLabel()
        self.lblDraglineId.setText("ID Dragline :")
        self.gridLayout1.addWidget(self.lblDraglineId, 1, 0)

        self.lblDate = QLabel()
        self.lblDate.setText("Date :")
        self.gridLayout1.addWidget(self.lblDate, 2, 0)

        self.date = QLabel()
        self.date.setText((str)(datetime.date.today()))
        self.gridLayout1.addWidget(self.date, 2, 1)

        self.lblTime = QLabel()
        self.lblTime.setText("Time :")
        self.gridLayout1.addWidget(self.lblTime, 3, 0)

        self.time = QLabel()
        self.gridLayout1.addWidget(self.time, 3, 1)

        self.table = QTableWidget()
        self.table.setColumnCount(3)  # Set three columns
        self.table.setRowCount(3)  # and three row

        # Set table headers
        self.table.setHorizontalHeaderLabels(["Volume (m3)", "Rendement Moyen (m3/h)", "Angle Moyenne (degree)"])
        self.table.setVerticalHeaderLabels(["Poste 3", "Poste 1", "Poste 2"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.hBoxDraglineInfo.addWidget(self.table)

        self.gridLayout2 = QGridLayout()
        self.hBoxDraglineInfo.addLayout(self.gridLayout2)

        self.lblDriverID = QLabel()
        self.lblDriverID.setText("ID Conducteur :")
        self.gridLayout2.addWidget(self.lblDriverID, 1, 0)

        self.driverID = QLabel()
        self.driverID.setText(str(self.dataConducteur[0]))
        self.gridLayout2.addWidget(self.driverID, 1, 1)

        self.lblDriverName = QLabel()
        self.lblDriverName.setText("Nom Conducteur :")
        self.gridLayout2.addWidget(self.lblDriverName, 2, 0)

        self.driverName = QLabel()
        self.driverName.setText(self.dataConducteur[1] + ' ' + self.dataConducteur[2])
        self.gridLayout2.addWidget(self.driverName, 2, 1)

        self.pushButton = QPushButton()
        self.pushButton.setText("Fermer")
        self.pushButton.setObjectName("btn")
        self.pushButton.setFixedHeight(30)
        self.pushButton.setFixedWidth(100)
        self.hBoxDraglineInfo.addWidget(self.pushButton)

        """Body Left"""

        self.hBoxBody = QHBoxLayout()

        self.widgetBody = QWidget()
        self.widgetBody.setObjectName("bodyWidget")
        self.widgetBody.setLayout(self.hBoxBody)

        self.vBox.addWidget(self.widgetBody, 50)

        self.vBoxBodyLeft = QVBoxLayout()
        self.hBoxBody.addLayout(self.vBoxBodyLeft)

        self.draglinePositionTitle = QLabel()
        self.draglinePositionTitle.setText("Position Dragline")
        self.draglinePositionTitle.setAlignment(QtCore.Qt.AlignCenter)
        self.draglinePositionTitle.setObjectName("positionTitle")
        self.vBoxBodyLeft.addWidget(self.draglinePositionTitle)

        self.layoutLeft = QVBoxLayout()
        self.vBoxBodyLeft.addLayout(self.layoutLeft)
        self.analogLeft = AnalogGaugeWidget()
        self.analogLeft.setFixedHeight(210)
        self.layoutLeft.addWidget(self.analogLeft)

        self.lcdCurrentAngleDragline = QLCDNumber()
        self.vBoxBodyLeft.addWidget(self.lcdCurrentAngleDragline)

        self.angleRestante = QLabel()
        self.angleRestante.setText("Angle Restante")
        self.angleRestante.setObjectName("qlabel")
        self.vBoxBodyLeft.addWidget(self.angleRestante)

        self.lcdAngleRestante = QLCDNumber()
        self.vBoxBodyLeft.addWidget(self.lcdAngleRestante)

        self.distanceRestante = QLabel()
        self.distanceRestante.setText("Distance Restante")
        self.distanceRestante.setObjectName("qlabel")
        self.vBoxBodyLeft.addWidget(self.distanceRestante)

        self.lcdDistanceRestante = QLCDNumber()
        self.vBoxBodyLeft.addWidget(self.lcdDistanceRestante)

        """Body Center"""

        self.vBoxBodyCenter = QVBoxLayout()
        self.hBoxBody.addLayout(self.vBoxBodyCenter)

        self.gridLayout3 = QGridLayout()
        self.vBoxBodyCenter.addLayout(self.gridLayout3)

        self.nombreGodets = QLabel()
        self.nombreGodets.setText("Nombre de godets")
        self.nombreGodets.setObjectName("label")
        self.gridLayout3.addWidget(self.nombreGodets, 1, 0)

        self.lcdNombreGodets = QLCDNumber()
        self.lcdNombreGodets.setFixedHeight(50)
        self.gridLayout3.addWidget(self.lcdNombreGodets, 2, 0)

        self.volume = QLabel()
        self.volume.setText("Volume Réalisé")
        self.gridLayout3.addWidget(self.volume, 1, 1)
        self.volume.setObjectName("label")

        self.lcdVolume = QLCDNumber()
        self.gridLayout3.addWidget(self.lcdVolume, 2, 1)

        self.rendementMoyen = QLabel()
        self.rendementMoyen.setText("Rendement Moyen")
        self.rendementMoyen.setAlignment(QtCore.Qt.AlignCenter)
        self.rendementMoyen.setObjectName("label")
        self.vBoxBodyCenter.addWidget(self.rendementMoyen)

        self.lcdRendementMoyen = QLCDNumber()
        self.lcdRendementMoyen.setFixedHeight(50)
        self.lcdRendementMoyen.setFixedWidth(400)
        self.vBoxBodyCenter.addWidget(self.lcdRendementMoyen, alignment=QtCore.Qt.AlignCenter)

        self.hBoxDistance = QHBoxLayout()
        self.vBoxBodyCenter.addLayout(self.hBoxDistance)

        """
        self.imgDistance = QLabel()
        self.pixmapDistanvce = QPixmap('../images/distance icon.png')
        self.pixmapSizeDistance = self.pixmapDistanvce.scaled(50, 50, Qt.KeepAspectRatio, Qt.FastTransformation)
        self.imgDistance.setPixmap(self.pixmapSizeDistance)

        self.hBoxDistance.addWidget(self.imgDistance)
        """

        self.spinBox = QSpinBox()
        self.spinBox.setPrefix('Entrer la distance realisée en m :   ')
        self.hBoxDistance.addWidget(self.spinBox)
        self.spinBox.setMinimum(0)
        self.spinBox.setMaximum(40)

        self.viewMap = QWebEngineView()
        self.viewMap.load(QtCore.QUrl('http://localhost:63342/Dragline/map/map.html'))
        self.vBoxBodyCenter.addWidget(self.viewMap)

        """Body Right"""

        self.vBoxBodyRight = QVBoxLayout()

        self.hBoxBody.addLayout(self.vBoxBodyRight)

        self.arrowPositionTitle = QLabel()
        self.arrowPositionTitle.setText("Position Flèche")
        self.arrowPositionTitle.setAlignment(QtCore.Qt.AlignCenter)
        self.arrowPositionTitle.setObjectName("positionTitle")
        self.vBoxBodyRight.addWidget(self.arrowPositionTitle)

        self.layoutRight = QVBoxLayout()
        self.vBoxBodyRight.addLayout(self.layoutRight)
        self.analogRight = AnalogGaugeWidget()
        self.analogRight.setFixedHeight(210)
        self.layoutRight.addWidget(self.analogRight)

        self.lcdCurrentAngleFleche = QLCDNumber()
        self.lcdCurrentAngleFleche.setObjectName("lcd")
        self.vBoxBodyRight.addWidget(self.lcdCurrentAngleFleche)

        self.angleMoyenne = QLabel()
        self.angleMoyenne.setText("Angle Moyenne")
        self.angleMoyenne.setObjectName("qlabel")
        self.vBoxBodyRight.addWidget(self.angleMoyenne)

        self.lcdAngleMoyenne = QLCDNumber()
        self.vBoxBodyRight.addWidget(self.lcdAngleMoyenne)

        """pour ajuster"""

        self.widgetRight = QWidget()
        self.vBoxBodyRight.addWidget(self.widgetRight)

        self.widgetRight1 = QWidget()
        self.vBoxBodyRight.addWidget(self.widgetRight1)

        """Connect"""
        self.pushButton.clicked.connect(self.close)
        # self.spinBox.valueChanged.connect(self.saveValueSpin)
        self.timerDAndAPosition = QTimer()
        self.timerDAndAPosition.setInterval(0)
        self.timerDAndAPosition.timeout.connect(self.showData)

        self.timerDR = QTimer()
        self.timerDR.setInterval(0)
        self.timerDR.timeout.connect(self.getDR)

        self.stop_flag_time = Event()
        self.getController = Controller(self.stop_flag_time)
        self.getController.start()
        self.getController.newTime.connect(self.updateTime)

    # ****************************************************************
    #    Methods
    # ***************************************************************
    
    def startTimer(self):
        self.timerDAndAPosition.start(3000)
        self.timerDR.start(3000)

    def showData(self):
        self.getDraglinePosition()
        self.getArrowPosition(self.poste)

    def getDR(self):
        wb = openpyxl.load_workbook(r'../files/Data IHM.xlsx')
        sheet = wb.active
        max_col = sheet.max_column
        min_col = sheet.min_column
        max_row = sheet.max_row
        min_row = sheet.min_row
        # Loop all columns name
        for i in range(min_col, max_col + 1):
            v = sheet.cell(row=min_row, column=i)
            m = str(v.value)
            if m == "D_R":
                for j in range(min_row + 1, max_row + 1):
                    cell = sheet.cell(row=j, column=i)
                    r = cell.value
                    if r is not None:
                        self.lcdDistanceRestante.display(str(r) + " pas")
                        QApplication.processEvents()
                        time.sleep(1)


    def getDraglinePosition(self):
        wb = openpyxl.load_workbook(r'../files/Data IHM.xlsx')
        sheet = wb.active
        max_col = sheet.max_column
        min_col = sheet.min_column
        max_row = sheet.max_row
        min_row = sheet.min_row
        # Loop all columns name
        for i in range(min_col, max_col + 1):
            v = sheet.cell(row=min_row, column=i)
            m = str(v.value)
            if m == "A_D":
                for j in range(min_row + 1, max_row + 1):
                    cell = sheet.cell(row=j, column=i)
                    r = cell.value
                    if r is not None:
                        self.lcdCurrentAngleDragline.display(r)
                        self.analogLeft.update_value(r)
                        if r < 90:
                            self.lcdCurrentAngleDragline.setStyleSheet("""QLCDNumber { 
                                                                            border-radius: 10px;
                                                                            background-color: rgb(142, 13, 20);
                                                                            color: white; }""")
                        else:
                            self.lcdCurrentAngleDragline.setStyleSheet("""QLCDNumber { 
                                                                            border-radius: 10px;
                                                                            background-color: rgb(57, 147, 15); 
                                                                            color: yellow;}""")
                        self.lcdAngleRestante.display(90 - r)
                        QApplication.processEvents()
                        time.sleep(1)


    def getArrowPosition(self, poste):
        wb = openpyxl.load_workbook(r'../files/Data IHM.xlsx')
        sheet = wb.active
        somme = 0
        som = 0
        max_col = sheet.max_column
        min_col = sheet.min_column
        max_row = sheet.max_row
        min_row = sheet.min_row
        # Loop all columns name
        for i in range(min_col, max_col + 1):
            v = sheet.cell(row=min_row, column=i)
            m = str(v.value)
            if m == "A_F":
                for k in range(min_col, max_col + 1):
                    l = sheet.cell(row=min_row, column=k)
                    s = str(l.value)
                    if s == "V_R":
                        for j in range(min_row + 1, max_row + 1):
                            cellAm = sheet.cell(row=j, column=i)
                            cellVr = sheet.cell(row=j, column=k)
                            a = cellAm.value
                            r = cellVr.value
                            if a is not None:
                                self.lcdCurrentAngleFleche.display(a)
                                self.analogRight.update_value(a)
                                if 70 < a < 110:
                                    self.lcdCurrentAngleFleche.setStyleSheet("""QLCDNumber { 
                                                            background-color: rgb(142, 13, 20);  
                                                            border-radius: 10px;
                                                            color: white; }""")
                                else:
                                    self.lcdCurrentAngleFleche.setStyleSheet("""QLCDNumber { 
                                                            border-radius: 10px;
                                                            background-color: rgb(57, 147, 15);
                                                            color: yellow; }""")
                                somme = somme + a
                                moy = somme / (k - 1)
                                self.lcdAngleMoyenne.display(moy)
                                self.itemAngleMoyenne = QTableWidgetItem(str(moy))
                                self.itemAngleMoyenne.setTextAlignment(Qt.AlignCenter)
                                self.itemAngleMoyenne.setFlags(Qt.ItemIsSelectable)
                                if poste == 3:
                                    self.table.setItem(0, 2, self.itemAngleMoyenne)
                                elif poste == 2:
                                    self.table.setItem(2, 2, self.itemAngleMoyenne)
                                else:
                                    self.table.setItem(1, 2, self.itemAngleMoyenne)

                            if r is not None:
                                self.lcdVolume.display(r)
                                self.lcdNombreGodets.display(j)
                                som = som + r
                                moyenne = som / (j - 1)
                                self.lcdRendementMoyen.display(moyenne)
                                self.itemVolume = QTableWidgetItem(str(r))
                                self.itemVolume.setTextAlignment(Qt.AlignCenter)
                                self.itemVolume.setFlags(Qt.ItemIsSelectable)
                                self.itemRendement = QTableWidgetItem(str(moyenne))
                                self.itemRendement.setTextAlignment(Qt.AlignCenter)
                                self.itemRendement.setFlags(Qt.ItemIsSelectable)
                                if poste == 3:
                                    self.table.setItem(0, 0, self.itemVolume)
                                    self.table.setItem(0, 1, self.itemRendement)
                                elif poste == 2:
                                    self.table.setItem(2, 0, self.itemVolume)
                                    self.table.setItem(2, 1, self.itemRendement)
                                else:
                                    self.table.setItem(1, 0, self.itemVolume)
                                    self.table.setItem(1, 1, self.itemRendement)

                            QApplication.processEvents()
                            time.sleep(1)


    def updateTime(self, timeInterval):
        self.time.setText(timeInterval)

    def close(self):
        self.saveValueSpin()

    def saveValueSpin(self):
        wb = openpyxl.load_workbook('../files/Data IHM.xlsx')
        sheet = wb.active
        max_col = sheet.max_column
        print(max_col)
        min_row = sheet.min_row
        print(min_row)
        if sheet.cell(row=min_row, column=max_col).value == 'Spin_Value':
            sheet.delete_cols(max_col)
            sheet.cell(row=min_row, column=max_col).value = 'Spin_Value'
            sheet.cell(column=max_col, row=min_row + 1, value=str(self.spinBox.value()))
            wb.save('../files/Data IHM.xlsx')

        else:
            sheet.cell(row=min_row, column=max_col + 1).value = 'Spin_Value'
            sheet.cell(column=max_col + 1, row=min_row + 1, value=str(self.spinBox.value()))
            wb.save('../files/Data IHM.xlsx')
