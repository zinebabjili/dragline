import folium
import openpyxl

from files.columnName import getNumColLongitude, getNumColLatitude

wb = openpyxl.load_workbook(r'../files/Data IHM.xlsx')
sheet = wb.active
max_row = sheet.max_row
min_row = sheet.min_row


#for j in range(min_row + 1, max_row + 1):
cellLongitude1 = sheet.cell(row=min_row + 1, column=getNumColLongitude("Lon_1"))
L1 = cellLongitude1.value
cellLatitude1 = sheet.cell(row=min_row + 1, column=getNumColLatitude("Lat_1"))
l1 = cellLatitude1.value
cellLongitude2 = sheet.cell(row=min_row + 1, column=getNumColLongitude("Lon_2"))
L2 = cellLongitude2.value
cellLatitude2 = sheet.cell(row=min_row + 1, column=getNumColLatitude("Lat_2"))
l2 = cellLatitude2.value
c = folium.Map(locations=[l1, L1])
if l1 != None and L1 != None :
    folium.Marker([l1, L1], icon=folium.Icon(color='green')).add_to(c)
    if l2 != None and L2 != None:
        folium.Marker([l2, L2], icon=folium.Icon(color='green')).add_to(c)
        #folium.PolyLine(locations=[[l1, L1], [l2, L2]], color='green').add_to(c)
        c.save('map.html')





#c = folium.Map(location=[l1, L1])
# a=data.readFile("Lat_1")
# b=data.readFile("Lon_1")
# c=data.readFile("Lat_2")
# d=data.readFile("Lon_2")
# c = folium.Map(location=[32.243, -7.9596])
# folium.Marker([32.243, -7.9596]).add_to(c)
# folium.Marker([33.367, -7.5732]).add_to(c)
# folium.PolyLine(locations=[[32.243, -7.9596], [33.367, -7.5732]], color='green').add_to(c)
# c.save('map.html')
# 0,3603603603603604
